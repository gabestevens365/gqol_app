import os
import re
import datetime
from datetime import date, datetime, timedelta
import mysql.connector
import pyodbc
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# Function to connect to the ADM/v5 MySQL database
def connect_to_v5_database():
    from credentials.credentials_sosdb import (
        adm_host,
        adm_port,
        adm_username,
        adm_password,
        adm_database,
    )

    try:
        mysql.connector.raise_on_warnings = True
        conn = mysql.connector.connect(
            host=adm_host,
            port=adm_port,
            user=adm_username,
            password=adm_password,
            database=adm_database,
        )
        print(f"Connected to the ADM/v5 (MySQL) database successfully!")
        return conn
    except mysql.connector.Error as error:
        print(f"Failed to connect to the ADM/v5 (MySQL) database: {error}")
        return None


# Function to connect to the Legacy MsSQL database
def connect_to_leg_database():
    from credentials.credentials_sosdb import (
        legacy_server,
        legacy_database,
        legacy_uid,
        legacy_pwd,
    )

    try:
        conn = pyodbc.connect(
            "DRIVER={SQL Server Native Client 11.0};"
            f"SERVER={legacy_server};"
            f"DATABASE={legacy_database};"
            f"UID={legacy_uid};"
            f"PWD={legacy_pwd};"
        )
        print(f"Connected to the Legacy (MsSQL) database successfully!")
        return conn
    except pyodbc.Error as error:
        print(f"Failed to connect to the Legacy (MsSQL) database: {error}")
        return None


# Function to connect to the AirVend MsSQL database
def connect_to_av_database():
    from credentials.credentials_sosdb import av_server, av_database, av_uid, av_pwd

    try:
        conn = pyodbc.connect(
            "DRIVER={SQL Server Native Client 11.0};"
            f"SERVER={av_server};"
            f"DATABASE={av_database};"
            f"UID={av_uid};"
            f"PWD={av_pwd};"
        )
        print(f"Connected to the AirVend (MsSQL) database successfully!")
        return conn
    except pyodbc.Error as error:
        print(f"Failed to connect to the AirVend (MsSQL) database: {error}")
        return None


# Function to test if an SSH tunnel is already established.
def check_ssh_tunnel(port):
    import psutil

    """Check if an SH tunnel is already established on the specified port."""
    established_connections = [
        conn for conn in psutil.net_connections() if conn.status == "LISTEN"
    ]
    for conn in established_connections:
        if conn.laddr.port == port:
            return True
    return False


# Function to connect to the CompanyKitchen MySQL database
def connect_to_ck_database():
    from credentials.credentials_sosdb import (
        ck_host,
        ck_port,
        ck_username,
        ck_password,
        ck_database,
    )

    try:
        # Check if the SSH tunnel is already active
        if not check_ssh_tunnel(3309):
            # Pause and prompt the user to connect to the SSH tunnel using PowerShell
            print(
                "Before connecting to the CompanyKitchen (MySQL) database, \
please open PowerShell and start the SSH tunnel on port 3309."
            )
            print("Once the tunnel is active, press 'C' to continue or 'Q' to quit.")

            while True:
                user_input = input().lower()
                if user_input == "c":
                    break
                elif user_input == "q":
                    print("Quitting the program.")
                    return None
                else:
                    print("Invalid input. Press 'C' to continue or 'Q' to quit.")

        # Try connecting to the CK database
        mysql.connector.raise_on_warnings = True
        conn = mysql.connector.connect(
            host=ck_host,
            port=ck_port,
            user=ck_username,
            password=ck_password,
            database=ck_database,
        )
        print(f"Connected to the CompanyKitchen (MySQL) database successfully!")
        return conn
    except mysql.connector.Error as error:
        print(f"Failed to connect to the CompanyKitchen (MySQL) database: {error}")
        return None


# Function to connect to the Avanti Azure Synapse database
def connect_to_avanti_database():
    from credentials.credentials_sosdb import (
        avanti_server,
        avanti_database,
        avanti_uid,
        avanti_pwd,
    )

    try:
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={avanti_server};"
            f"DATABASE={avanti_database};"
            "Authentication=ActiveDirectoryPassword;"
            f"UID={avanti_uid};"
            f"PWD={avanti_pwd};"
        )
        print(f"Connected to the Avanti (MsSQL / Azure Synapse) database successfully!")
        return conn
    except pyodbc.Error as error:
        print(
            f"Failed to connect to the Avanti (MsSQL / Azure Synapse) database: {error}"
        )
        return None


# Function to execute a SQL query and return the results as a pandas DataFrame
def execute_query(conn, query_file):
    query_starttime = datetime.now()
    print(
        f"\nStarting the {query_file} query at {query_starttime.strftime('%Y-%m-%d @ %H:%M:%S')}"
    )
    try:
        cursor = conn.cursor()

        with open(query_file, "r") as file:
            query = file.read()

        cursor.execute(query)
        results = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        cursor.close()

        # Convert bytes columns to utf-8 strings
        decoded_results = []
        for row in results:
            decoded_row = []
            for item in row:
                if isinstance(item, bytes):
                    decoded_row.append(item.decode("utf-8"))
                else:
                    decoded_row.append(item)
            decoded_results.append(decoded_row)

        df = pd.DataFrame(decoded_results, columns=columns)
        query_endtime = datetime.now()
        query_elapsed = (query_endtime - query_starttime).total_seconds()
        print(
            f"Ending the {query_file} query at {query_endtime.strftime('%Y-%m-%d @ %H:%M:%S')}"
            f", elapsed time: {int(query_elapsed)} seconds."
        )
        return df
    except mysql.connector.Error as error:
        print(f"Failed to execute the {query_file} query: {error}\n")
        return None


# Function to format the worksheet
def format_worksheet(df, workbook, worksheet):
    format_gray = workbook.add_format({"pattern": 18, "bg_color": "#CCCCCC"})
    format_red = workbook.add_format({"bg_color": "#FFB6C1"})
    format_orange = workbook.add_format({"bg_color": "#FFE4B5"})
    format_yellow = workbook.add_format({"bg_color": "#FFFFE0"})
    format_green = workbook.add_format({"bg_color": "#BDFCC9"})
    format_blue = workbook.add_format({"bg_color": "#BFEFFF"})
    format_purple = workbook.add_format({"bg_color": "#E6E6FA"})

    formats = [
        format_red,
        format_orange,
        format_yellow,
        format_green,
        format_blue,
        format_purple,
    ]

    info_cols = [
        get_column_letter(i + 1)
        for i, col in enumerate(df.columns)
        if col.endswith(" Info")
    ]
    last_col = get_column_letter(df.shape[1])

    # Apply formatting to the first column, and all the info_cols
    worksheet.set_column("A:A", None, format_gray)
    for col_letter in info_cols:
        worksheet.set_column(f"{col_letter}:{col_letter}", None, format_gray)

    # Apply additional formatting to the header row
    start_col = "A"
    for i, info_col in enumerate(info_cols):
        end_col = get_column_letter(column_index_from_string(info_col) - 1)
        worksheet.conditional_format(
            f"{start_col}1:{end_col}1",
            {"type": "no_blanks", "format": formats[i % len(formats)]},
        )
        start_col = info_col
    # Apply conditional formatting from the last " Info" column to the last column in the dataframe
    worksheet.conditional_format(
        f"{start_col}1:{last_col}1",
        {"type": "no_blanks", "format": formats[len(info_cols) % len(formats)]},
    )


# Temporary function with my cleanup tasks
def v5_adm_alldev_cleanup(df):
    # Fix a couple columns of known bad data.
    try:
        condition = df["Operator Name"] == "CanteenCanada"
        count = df.loc[condition, "Operator Group"].shape[0]
        df.loc[condition, "Operator Group"] = "Canteen Canada"
        print(f"'Canteen Canada' updated in {count} Operator Group rows")
    except Exception as e:
        print("Error updating 'Operator Group' for 'Canteen Canada':", e)

    try:
        condition = df["Operator Group"] == "Canteen_Dining"
        count = df.loc[condition, "Operator Group"].shape[0]
        df.loc[condition, "Operator Group"] = "CanteenDining"
        print(f"'Canteen_Dining' updated in {count} Operator Group rows")
    except Exception as e:
        print("Error updating 'Operator Group' for 'Canteen_Dining':", e)

    try:
        condition = (df["Device Serial"].str.startswith("bea")) | (
            df["Device Serial"].str.startswith("365P")
        )
        count = df.loc[condition, "Device Type"].shape[0]
        count2 = df.loc[condition, "Device HW Type"].shape[0]
        df.loc[condition, "Device Type"] = "BEACON"
        df.loc[condition, "Device HW Type"] = "Beacon"
        print(
            f"'Beacons' updated in {count} Device Type rows and {count2} Device HW Type"
        )
    except Exception as e:
        print(
            "Error updating 'Device Type' and 'Device HW Type' for 'bea' or '365P':", e
        )

    try:
        condition = df["Device Serial"].str.startswith("vsh")
        count = df.loc[condition, "Device HW Type"].shape[0]
        df.loc[condition, "Device HW Type"] = "Gen3"
        print(f"'Gen3' updated in {count} Device HW Type rows")
    except Exception as e:
        print("Error updating 'Device HW Type' for 'vsh':", e)

    try:
        condition = df["Device Serial"].str.startswith("ksk")
        count = df.loc[condition, "Device HW Type"].shape[0]
        df.loc[condition, "Device HW Type"] = "ReadyTouch(Unknown)"
        print(f"'ReadyTouch(Unknown) updated in {count} Device HW Type rows")
    except Exception as e:
        print("Error updating 'Device HW Type' for 'ksk':", e)

    # try:
    # condition = df["Device Serial"].str.startswith("SWL")
    # count = df.loc[condition, "Device HW Type"].shape[0]
    # df.loc[condition, "Device HW Type"] = "Pico-Stockwell"
    # print(f"'Pico-Stockwell' updated in {count} Device HW Type rows")
    # except Exception as e:
    # print("Error updating 'Device HW Type' for 'SWL':", e)

    return df


def get_os_version_column_name(df, possible_column_names):
    for column_name in possible_column_names:
        if column_name in df.columns:
            return column_name
    return None


def excel_date_to_datetime(excel_date):
    # Excel's serial date format starts on 1900-01-01
    if pd.isnull(excel_date):
        return None
    try:
        return datetime(1899, 12, 30) + timedelta(days=int(excel_date))
    except ValueError:
        return None


def add_eos_dates(df, eos_dates_df, possible_os_version_columns):
    print("Adding EOS Dates")
    # Infer platform from the "Platform" column of the Dataframe.
    if "Platform" not in df.columns:
        print("The 'Platform' column is missing from the input Dataframe.")
        return df
    unique_platforms = df["Platform"].unique()
    if len(unique_platforms) > 1:
        print(
            f"Multiple platforms found in the DataFrame: {unique_platforms}. Ensure the dataframe only contains one."
        )
        return df
    platform_name = unique_platforms[0]

    # Figure out the OS Ver column name for this specific dataset.
    os_version_col = get_os_version_column_name(df, possible_os_version_columns)
    if os_version_col is None:
        print(
            f"None of the specified OS version columns found in DataFrame. Available columns: {df.columns}"
        )
        return df

    # Standardize data formatting for merging
    df[os_version_col] = df[os_version_col].str.strip().str.lower()
    eos_dates_df["OS & Ver"] = eos_dates_df["OS & Ver"].str.strip().str.lower()

    # Rename 'Platform' column in eos_dates_df to 'Platform2' before the merge
    eos_dates_df_renamed = eos_dates_df.rename(columns={"Platform": "Platform2"})

    # Filter eos_dates_df_renamed for the current platform
    platform_eos_dates = eos_dates_df_renamed[
        eos_dates_df_renamed["Platform2"] == platform_name
    ]

    if platform_eos_dates.empty:
        print(f"No EOS data found for platform: {platform_name}")
        return df

    # Before merging, ensure EOS dates are in datetime format
    eos_dates_df_renamed["EOS Date"] = eos_dates_df_renamed["EOS Date"].apply(
        excel_date_to_datetime
    )

    # Merge the device DataFrame with the platform specific EOS dates
    merged_df = pd.merge(
        df,
        platform_eos_dates,
        left_on=os_version_col,
        right_on="OS & Ver",
        how="left",
    )
    # Drop the 'Platform' and 'OS & Ver' columns
    columns_to_drop = [
        col for col in ["Platform2", "OS & Ver"] if col in merged_df.columns
    ]
    merged_df.drop(columns_to_drop, axis=1, inplace=True, errors="ignore")

    return merged_df


# Function to save a DataFrame as an Excel file
def save_to_excel(dataframes, file_name, sheet_names):
    try:
        print("Formatting and Saving. Please wait...")
        try:
            if len(dataframes) != len(sheet_names):
                raise ValueError("Number of dataframes and sheet names must be equal.")

            with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
                for df, sheet_name in zip(dataframes, sheet_names):
                    # Write DataFrames to Excel
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]

                    # Apply Column Filters & Freeze the top row
                    last_col = len(df.columns) - 1
                    worksheet.autofilter(0, 0, len(df), last_col)
                    worksheet.freeze_panes(1, 0)

                    # Auto-width columns
                    for i, col in enumerate(df.columns):
                        column_len = max(df[col].astype(str).str.len().max(), len(col))
                        column_len = min(column_len, 59)  # limit column width to 59
                        worksheet.set_column(i, i, column_len)

                    # Format specific sheets
                    if sheet_name == "All ADM-v5 Devices":
                        format_worksheet(df, workbook, worksheet)
                    elif sheet_name == "All Legacy Devices":
                        format_worksheet(df, workbook, worksheet)
                    elif sheet_name == "All AirVend Devices":
                        format_worksheet(df, workbook, worksheet)
                    elif sheet_name == "All CompanyKitchen Devices":
                        format_worksheet(df, workbook, worksheet)
                    elif sheet_name == "All Avanti Devices":
                        format_worksheet(df, workbook, worksheet)

            print(f"Data saved to '{file_name}' successfully!")
            main_menu()
        except Exception as error:
            print(f"Failed to save the data to Excel file: {error}")
            main_menu()
    except Exception as error:
        print("Error details:", error)
        raise  # Re-raise the exception for full traceback


# Find the most recent report in the folder
def find_latest_report(report_path, report_name, report_date):
    print(
        f"Finding the latest report with.\nPath: {report_path}\nName: {report_name}\nDate: {report_date}"
    )
    # Filter files in the report_path that start with report_name and include report_date
    matching_files = [
        file
        for file in os.listdir(report_path)
        if file.startswith(report_name) and report_date in file
    ]

    if matching_files:
        # Sort the matching_files based on their modification time (newest first)
        sorted_files = sorted(
            matching_files,
            key=lambda f: os.path.getmtime(os.path.join(report_path, f)),
            reverse=True,
        )
        latest_report = sorted_files[0]
        return latest_report
    else:
        return None


# This really should be combined with find_latest_report
def find_latest_csv(report_path, report_prefix):
    report_date = date.today().strftime(
        "%Y-%m-%d"
    )  # Assuming the report is named with today's date
    matching_files = [
        file
        for file in os.listdir(report_path)
        if file.startswith(report_prefix) and report_date in file
    ]
    if matching_files:
        # Assuming these files are named with a consistent date format
        latest_file = sorted(matching_files)[-1]  # Sorting to get the latest
        return latest_file
    else:
        return None


def extract_date_from_filename(filename):
    # Use a regular expression to find a date pattern in the filename
    # This pattern matches dates in the format YYYY-MM-DD, optionally followed by _hhmm
    date_pattern = r"(\d{4}-\d{2}-\d{2})"
    match = re.search(date_pattern, filename)
    if match:
        return match.group(1)  # Return the matched date part
    else:
        return None  # Or raise an error/exception if appropriate


def create_variation_for_operator_group_xlsxwriter(
    latest_report, group_names, variation_name
):
    report_path = "./reports/AllDevices/"
    variation_filename = f"{report_path}All-Devices_Report_{datetime.now().strftime('%Y-%m-%d')}_{variation_name}.xlsx"

    wb = pd.ExcelFile(latest_report)  # Use pandas to read the Excel file

    with pd.ExcelWriter(
        variation_filename, engine="xlsxwriter"
    ) as writer:  # Open the writer
        for sheet_name in wb.sheet_names:
            df = wb.parse(sheet_name)  # Read each sheet into a DataFrame

            # Filter based on 'Operator Group'
            filtered_df = df[df["Operator Group"].isin(group_names)].copy()

            if not filtered_df.empty:
                filtered_df.to_excel(
                    writer, sheet_name=sheet_name, index=False
                )  # Write filtered DataFrame to a sheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Example of formatting: auto-width columns, apply filter
                for column in filtered_df.columns:
                    column_width = max(
                        filtered_df[column].astype(str).apply(len).max(), len(column)
                    )
                    col_idx = filtered_df.columns.get_loc(column)
                    worksheet.set_column(
                        col_idx, col_idx, column_width + 2
                    )  # Adjusting column width

                worksheet.autofilter(
                    0, 0, 0, len(filtered_df.columns) - 1
                )  # Apply autofilter
                worksheet.freeze_panes(1, 0)  # Freeze top row

                # Format specific sheets
                if sheet_name == "All ADM-v5 Devices":
                    format_worksheet(filtered_df, workbook, worksheet)
                elif sheet_name == "All Legacy Devices":
                    format_worksheet(filtered_df, workbook, worksheet)
                elif sheet_name == "All AirVend Devices":
                    format_worksheet(filtered_df, workbook, worksheet)
                elif sheet_name == "All CompanyKitchen Devices":
                    format_worksheet(filtered_df, workbook, worksheet)
                elif sheet_name == "All Avanti Devices":
                    format_worksheet(filtered_df, workbook, worksheet)

            else:
                print(
                    f"No data after filtering for '{sheet_name}' in {variation_name}, skipping this sheet."
                )

        # The workbook is saved when exiting the 'with' context

    print(f"Variation '{variation_name}' saved as {variation_filename}")


def finalize_all_devices_report():
    # Split the All_Devices_Report into several different files based on the requesting customer.
    # Open the recently saved All_Devices_Report.
    print("Finalizing the All_Devices_Report.")
    formatted_date = date.today().strftime("%Y-%m-%d")
    formatted_time = datetime.now().strftime("%H%M")
    report_path = "./reports/AllDevices/"
    report_name = "All_Devices_Report"

    latest_report = report_path + find_latest_report(
        report_path, report_name, formatted_date
    )
    print(f"Found {latest_report}")

    # TODO: Check to see if OperatorGroup-Specific reports have already been generated for that file, based on
    #  filenames. All_Devices_Report_{date}.xlsx is the default, the variation is
    #  All_Devices_Report_{date}_OperatorGroup.xlsx. If they already exist, say so and return to the menu, if not,
    #  create them.

    # Create the "CompassGroup" variation.
    compass_groups = [
        "Canteen",
        "Canteen Canada",
        "Canteen_OCS",
        "CanteenDining",
        "CompassGroup",
        "Eurest",
        "Canteen FP",
        "Dining - Canteen",
    ]
    create_variation_for_operator_group_xlsxwriter(
        latest_report, compass_groups, "CompassGroup"
    )

    # Create the "CFG" variation.
    cfg_group = ["CanteenFranchiseGroup", "FiveStar", "CFG", "CFG QC"]
    create_variation_for_operator_group_xlsxwriter(latest_report, cfg_group, "CFG")

    # Create the "CanteenCanada" variation.
    canada_group = ["Canteen Canada"]
    create_variation_for_operator_group_xlsxwriter(
        latest_report, canada_group, "CanteenCanada"
    )


# Main menu function
def main_menu():
    print("=== Main Menu ===")
    print("1. All Devices Report")
    print("2. Kiosk Age Report")
    print("3. Finalize All-Devices Report")
    print("0. Quit")

    choice = input("Enter your choice: ")

    if choice == "1":
        report_builder("./reports/AllDevices/", "All_Devices_Report")
    elif choice == "2":
        report_builder("./reports/KioskAge/", "KioskAge_Report")
    elif choice == "3":
        finalize_all_devices_report()
    elif choice == "0":
        print("Exiting the program. Goodbye!")
        return
    else:
        print("Invalid choice. Please try again.")
        main_menu()


def generate_report_name(report_path, report_name):
    formatted_date = date.today().strftime("%Y-%m-%d")
    formatted_time = datetime.now().strftime("%H%M")
    return f"{report_path}{report_name}_{formatted_date}_{formatted_time}.xlsx"


def generate_report(final_file_name, report_name):
    if report_name == "All_Devices_Report":
        alldevice_report_365rm_writer(final_file_name)
    elif report_name == "KioskAge_Report":
        kiosk_age_report_writer(final_file_name)
    else:
        print("There is a problem in the Report Name.")
        main_menu()


def report_builder(report_path, report_name):
    formatted_date = date.today().strftime("%Y-%m-%d")
    formatted_time = datetime.now().strftime("%H%M")

    print(f"{report_name} for {formatted_date} at {formatted_time}...")

    final_file_name = generate_report_name(report_path, report_name)

    report_exists = find_latest_report(report_path, report_name, formatted_date)

    if report_exists is None:
        print(f"Generating the first copy of {report_name} for {formatted_date}.")
        generate_report(final_file_name, report_name)
        return

    print(f"{report_path}{report_exists} already exists!\n")
    choice = input("Do you want to generate a new copy? [Y/N]: ").strip().upper()

    if choice == "Y":
        print(f"Generating a new copy of {report_name} for {formatted_date}.")
        generate_report(final_file_name, report_name)
    elif choice == "N":
        main_menu()
    else:
        print("Invalid Choice.")
        main_menu()


def map_operator_group(source_df, mapping_df):
    operator_group_map = mapping_df.set_index("Operator Name")[
        "Operator Group"
    ].to_dict()

    source_df["Operator Group"] = source_df["Operator Name"].map(operator_group_map)

    return source_df


def execute_and_save_queries(db_info, filename):
    result_dfs = []
    sheet_names = []

    # Get the End-of-Support dates for our various OS / Platform combinations.
    eos_dates_df = pd.read_csv(
        "C:/Users/gabe.stevens/PycharmProjects/gqol_app/queries/OS_EOS_dates.csv"
    )

    # Get the Org-Group Mappings for Avanti and CK
    ogmap_avanti_df = pd.read_csv(
        "C:/Users/gabe.stevens/PycharmProjects/gqol_app/queries/OrgGroupMapping_Avanti.csv"
    )
    ogmap_ck_df = pd.read_csv(
        "C:/Users/gabe.stevens/PycharmProjects/gqol_app/queries/OrgGroupMapping_CK.csv"
    )

    # The Avanti data needs to be pulled manually, via two separate CSV files. This section is to get that data.
    today_str = datetime.now().strftime("%Y-%m-%d")
    # Get the AMS_devices file
    ams_devices_df = pd.read_csv(
        f"C:/Users/gabe.stevens/PycharmProjects/gqol_app/queries/dailysheets/AMS_devices_{today_str}.csv",
        dtype={"Op Zip": str},
    )
    if ams_devices_df is not None:
        print(f"Successfully loaded AMS_devices_{today_str}.csv")

    bomgar_df = pd.read_csv(
        f"C:/Users/gabe.stevens/PycharmProjects/gqol_app/queries/dailysheets/Bomgar_{today_str}.csv",
        usecols=["Name", "Connection Status", "Operating System"],
        skiprows=1,
    )
    # Drop rows where 'Name' is blank (it's the required row)
    bomgar_df = bomgar_df.dropna(subset=["Name"])
    # Trim the Connection Status to only the timestamp.
    bomgar_df["Extracted Datetime"] = pd.to_datetime(
        bomgar_df["Connection Status"].str.extract(
            "(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2} [AP]M)"
        )[0]
    )
    # Sort by Name and then Extracted Datetime to ensure the latest status is last
    bomgar_df = bomgar_df.sort_values(by=["Name", "Extracted Datetime"])
    # Deduplicate keeping the last entry for each Name, which is the most recent
    bomgar_df = bomgar_df.drop_duplicates(subset=["Name"], keep="last")
    # Map the Operating System from Bomgar into the AMS 'OS Version' column
    os_version_map = bomgar_df.set_index("Name")["Operating System"]
    ams_devices_df["OS Version"] = ams_devices_df["Kiosk Name"].map(os_version_map)
    ams_devices_df = add_eos_dates(ams_devices_df, eos_dates_df, ["OS Version"])
    # End the Avanti section.

    # Start the loop to get the SQL results
    for info in db_info:
        connection = info["connect_func"]()
        if connection is None:
            print(f"Failed to connect to {info['name']} database.")
            continue  # Skip this database

        result_df = execute_query(connection, info["query_file"])

        if result_df is None:
            print(f"Query failed for {info['name']} database.")
            connection.close()
            continue  # Skip saving this result

        if info["query_file"] == "./queries/All-Devices_CK.sql":
            result_df = map_operator_group(result_df, ogmap_ck_df)
        elif info["query_file"] == "./queries/All-Devices_AV.sql":
            # Apply manual rule for 'Operator Name' containing 'Canteen'
            exceptions = [
                "Little Canteen, LLC",
                "Mama's Canteen Vending",
                "SE Canteen Memphis",
            ]
            # Use a mask to identify rows that meet the condition
            mask = (result_df["Operator Name"].str.contains("Canteen")) & (
                ~result_df["Operator Name"].isin(exceptions)
            )
            # Update 'Operator Group' for rows matching the condition
            result_df.loc[mask, "Operator Group"] = "Canteen"

        possible_os_version_columns = ["Kiosk OS Version", "OS Version"]
        result_df = add_eos_dates(result_df, eos_dates_df, possible_os_version_columns)
        result_df = v5_adm_alldev_cleanup(result_df)
        result_dfs.append(result_df)
        sheet_names.append(info["sheet_name"])
        connection.close()

    # Add the Avanti results into the final results_dfs and sheet_names
    ams_devices_df = map_operator_group(ams_devices_df, ogmap_avanti_df)
    result_dfs.append(ams_devices_df)
    sheet_names.append("All Avanti Devices")

    if result_dfs:
        save_to_excel(result_dfs, filename, sheet_names)


def alldevice_report_365rm_writer(filename):
    db_info = [
        {
            "name": "v5",
            "connect_func": connect_to_v5_database,
            "query_file": "./queries/All-Devices_v5.sql",
            "sheet_name": "All ADM-v5 Devices",
        },
        {
            "name": "Legacy",
            "connect_func": connect_to_leg_database,
            "query_file": "./queries/All-Devices_Legacy.sql",
            "sheet_name": "All Legacy Devices",
        },
        {
            "name": "AirVend",
            "connect_func": connect_to_av_database,
            "query_file": "./queries/All-Devices_AV.sql",
            "sheet_name": "All AirVend Devices",
        },
        {
            "name": "CompanyKitchen",
            "connect_func": connect_to_ck_database,
            "query_file": "./queries/All-Devices_CK.sql",
            "sheet_name": "All CompanyKitchen Devices",
        },
        {
            "name": "Avanti",
            "connect_func": connect_to_avanti_database,
            "query_file": "./queries/All-Devices_Avanti.sql",
            "sheet_name": "All Avanti Devices",
        },
    ]

    execute_and_save_queries(db_info, filename)


# KioskAge Report
def kiosk_age_report_writer(filename):
    # Get the SQL query
    query_file_v5 = "./queries/KioskAge_v5_Report.sql"
    query_file_rt = "./queries/KioskAge_RT_Report.sql"

    # Connect to the v5 / RT database
    connection_v5 = connect_to_v5_database()
    if connection_v5 is None:
        return

    # Execute the v5 query
    result_v5_df = execute_query(connection_v5, query_file_v5)
    if result_v5_df is None:
        connection_v5.close()
        return

    # Execute the ReadyTouch query
    result_rt_df = execute_query(connection_v5, query_file_rt)
    if result_rt_df is None:
        connection_v5.close()
        return

    # Exclude specific CPU Products from the dataframe
    excluded_v5_cpu_products = [
        "Elo AiO",
        "Elo AiO X3",
        "EloPOS E2/S2/H2",
        "EloPOS E3/S3/H3",
        "MMH81AP-FH",
        "OptiPlex 7010",
        "S11G",
        "S11M",
        "W11G",
        "W11HS2",
    ]
    excluded_rt_cpu_products = [
        "S11G",
        "W8LPL",
        "EloPOS E3/S3/H3",
        "EloPOS E2/S2/H2",
        "To Be Filled By O.E.M.",
    ]
    result_v5_df = result_v5_df[
        ~result_v5_df["CPU Product"].isin(excluded_v5_cpu_products)
    ]
    result_rt_df = result_rt_df[
        ~result_rt_df["CPU Product"].isin(excluded_rt_cpu_products)
    ]

    # v5 Report -- Add the new columns "Sage Go-Live", "Device Age", and "Resolution Path"
    (
        result_v5_df["Sage Go-Live"],
        result_v5_df["Device Age"],
        result_v5_df["Resolution Path"],
    ) = ("", "", "")

    # Import the Sage Data and convert Excel dates to datetime objects, handling NaN values
    csv_file_path = "./queries/SageData_v5_golives.csv"
    sage_data_df = pd.read_csv(csv_file_path)

    # Convert dates to Excel date format
    sage_data_df["WentLiveOn"] = pd.to_datetime(
        sage_data_df["WentLiveOn"], format="%m/%d/%Y"
    )
    excel_date_start = datetime(1899, 12, 30)  # Excel's serial date start
    sage_data_df["WentLiveOn"] = (sage_data_df["WentLiveOn"] - excel_date_start).dt.days

    # Merge the result_v5_df and sage_data_df on the 'Device Serial' and 'SerialNumber' columns
    merged_df = pd.merge(
        result_v5_df,
        sage_data_df[["SerialNumber", "WentLiveOn"]],
        how="left",
        left_on="Device Serial",
        right_on="SerialNumber",
    )

    # Retrieve only the device serials that have a valid 'WentLiveOn' value in 'merged_df'
    valid_device_serials = merged_df.loc[
        merged_df["WentLiveOn"].notnull(), "Device Serial"
    ].values.tolist()
    valid_went_live_on = merged_df.loc[
        merged_df["WentLiveOn"].notnull(), "WentLiveOn"
    ].values.tolist()

    # Create a dictionary mapping valid device serials to their corresponding WentLiveOn values
    valid_mapping = dict(zip(valid_device_serials, valid_went_live_on))

    # Safely update 'Sage Go-Live' in 'result_v5_df'
    result_v5_df["Sage Go-Live"] = result_v5_df["Device Serial"].apply(
        lambda x: valid_mapping.get(x, "")
    )

    # Calculate the Device Age based on Device Go-Live and Sage Go-Live
    result_v5_df["Device Age"] = None
    device_go_live = pd.to_datetime(
        result_v5_df["Device Go-Live"], format="%m/%d/%Y", errors="coerce"
    )
    sage_go_live = pd.to_datetime(
        result_v5_df["Sage Go-Live"], format="%m/%d/%Y", errors="coerce"
    )
    max_go_live_dates = pd.concat([device_go_live, sage_go_live], axis=1).max(axis=1)
    current_date = pd.Timestamp.now()
    result_v5_df["Device Age"] = (current_date - max_go_live_dates).dt.days // 365

    # ReadyTouch Report -- Add a new column "Path Forward" and fill it with relevant values.
    result_rt_df["Resolution Path"] = ""
    result_rt_df.loc[
        result_rt_df["OS Version"] == "Ubuntu 14.04", "Resolution Path"
    ] = "Upgrade (Ubuntu 14.04)"
    result_rt_df.loc[
        result_rt_df["CPU Product"].isnull() | (result_rt_df["CPU Product"] == ""),
        "Resolution Path",
    ] = "Investigate (Not in Dash)"
    result_rt_df.loc[
        result_rt_df["CPU Product"].str.contains("^Opti", na=False), "Resolution Path"
    ] = "Replace (CPU Not Eligible)"
    result_rt_df.loc[
        result_rt_df["OS Version"] == "Ubuntu 20.04", "Resolution Path"
    ] = "Up-to-Date (Ubuntu 20.04)"

    # v5 Report -- Fill the "Resolution Path" column with relevant values
    four_months_ago = pd.Timestamp.now() - pd.DateOffset(months=4)
    result_v5_df["Resolution Path"] = "Investigate"
    result_v5_df.loc[
        result_v5_df["OS Version"].notnull()
        & result_v5_df["OS Version"].str.startswith("Cent"),
        "Resolution Path",
    ] = "Upgrade (CentOS)"
    result_v5_df.loc[
        result_v5_df["OS Version"] == "Ubuntu 14.04", "Resolution Path"
    ] = "Upgrade (Ubuntu 14)"
    result_v5_df.loc[
        result_v5_df["Device Age"] >= 6, "Resolution Path"
    ] = "Replace (6+ years old)"
    mask = (result_v5_df["OS Version"].str.startswith("Cent")) & (
        result_v5_df["CPU Product"].str.startswith("W10")
    )
    result_v5_df.loc[mask, "Resolution Path"] = "Replace (CentOS on W10* cpu)"
    result_v5_df.loc[
        result_v5_df["Device Serial"].str.startswith("VSH310"), "Resolution Path"
    ] = "Replace (VSH310xxx)"
    mask = result_v5_df["Device Serial"].str.startswith(("VSH1", "VSH2"))
    result_v5_df.loc[mask, "Resolution Path"] = "Replace (VSH1 / VSH2)"
    mask = (
        (result_v5_df["Device Serial"].str.startswith("VSH3"))
        & (result_v5_df["Location Name"].isin(["", "Orphan Loc"]))
        & (result_v5_df["Device Last Sync"] < four_months_ago)
    )
    result_v5_df.loc[mask, "Resolution Path"] = "Decommission (VSH3 unused orphan)"
    mask = (result_v5_df["Device Serial"].str.startswith(("VSH1", "VSH2"))) & (
        result_v5_df["Location Name"].isin(["", "Orphan Loc"])
    )
    result_v5_df.loc[mask, "Resolution Path"] = "Decommission (VSH1 / VSH2 orphan)"
    result_v5_df.loc[
        result_v5_df["OS Version"] == "Ubuntu 20.04", "Resolution Path"
    ] = "Already Up-to-Date"

    # Extra Cleanup
    result_v5_df.loc[
        result_v5_df["Operation Name"] == "CanteenCanada", "Operation Group"
    ] = "Canteen Canada"
    result_v5_df.loc[
        result_v5_df["Operation Group"] == "Canteen_Dining", "Operation Group"
    ] = "CanteenDining"
    #    result_rt_df.loc[result_rt_df]
    # Save the result to an Excel file
    sheet1 = "v5 KioskAges"
    sheet2 = "RT KioskAges"
    save_to_excel([result_v5_df, result_rt_df], filename, [sheet1, sheet2])


# Entry point of the program
if __name__ == "__main__":
    main_menu()
